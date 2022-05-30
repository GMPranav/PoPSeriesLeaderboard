import re
import json
import numpy
import collections
import openpyxl as xl

runners_wb = xl.load_workbook('Leaderboard.xlsx')
runners_sheet = runners_wb.active

with open('test.npy', 'rb') as f:
    runsArray = numpy.load(f, allow_pickle=True)

pop1anyDOS = []
pop1anyDOSNMG = []
pop1anyNES = []
pop1anyAppleII = []
pop1anyTG16 = []
pop1anySMS = []
pop1anyGBC = []
pop1anyAmiga = []
pop1anyMD = []
pop1anyGenesis = []
pop1anyLS = []
pop1anyLSNMG = []
pop1hundo = []
pop1hundoNMG = []
pop1pacifist = []
pop1istaria = []
pop1RPR = []
pop1SNESanyNW = []
pop1SNESany = []
pop1SNESanyNGS = []
pop1SNESSR = []
pop1SNES30AP = []
pop1SNESTQoL = []
pop1SNESEvol = []
pop1SNESTDC = []
pop2any = []
pop3dany = []
pop3danyAT = []
popanany = []
popsotACZip = []
popsotACZipl = []
popsotACNMG = []
popsotanyZip = []
popsotanyZipl = []
popsotanyNMG = []
popsothundoZip = []
popsothundoNMG = []
popsotGBAany = []
popwwanyZip = []
popwwanyZipl = []
popwwanyNMG = []
popwwTEZip = []
popwwTEZipl = []
popwwTENMG = []
popwwACNMG = []
popwwAChZipl = []
popwwAChNMG = []
popwwAAZip = []
popt2tanyZip = []
popt2tanyZipl = []
popt2tanyNMG = []
popt2tAPZip = []
popt2tAPZipl = []
popt2tAPNMG = []
popt2tPSPUSNMG = []
popclassicany = []
pop2k8anySt = []
pop2k8anyNMG = []
pop2k8AS = []
pop2k8Epilougeany = []
pop2k8Hacked = []
pop2k8anyDupe = []
poptfkanyCon = []
poptfkanyEmu = []
poptfsanySt = []
poptfsanyNEG = []
poptfsanyNMG = []
poptfsPSPany = []
poptfsWiianyNG = []
poptfsWiianyNGP = []
poptfsDSanyCon = []
poptfsDSanyEmu = []
mobileHA = []
mobileSnFR = []
mobileSoT = []
mobileWW = []
mobileT2T = []
mobileClassic = []
mobile2k8 = []
mobileTFS = []
multiSTanyZip = []
multiSTanyZipl = []
multiSTanyNMG = []
multiSTCompZip = []
multiSTCompZipl = []
multiSTCompNMG = []
runnersArray = {"":0}

for i in runsArray:
    if(i['category'] == 'rklx8pnk'):
        if(i['subcat'] == 'rqvvyeyq'):
            pop1anyDOS.append(i)
        if(i['subcat'] == '4qyxp34l'):
            pop1anyDOSNMG.append(i)
        if(i['subcat'] == '5le2xv6l'):
            pop1anyNES.append(i)
        if(i['subcat'] == '81w2ve51'):
            pop1anyAppleII.append(i)
        if(i['subcat'] == '81wm0j5q'):
            pop1anyTG16.append(i)
        if(i['subcat'] == 'zqo50dpl'):
            pop1anyGBC.append(i)
        if(i['subcat'] == 'p12z2m41'):
            pop1anySMS.append(i)
        if(i['subcat'] == '81pdo5v1'):
            pop1anyAmiga.append(i)
        if(i['subcat'] == '810y2o5l'):
            pop1anyMD.append(i)
        if(i['subcat'] == 'klrv7woq'):
            pop1anyGenesis.append(i)
    
    if(i['category'] == 'ndx9xq5d'):
        if(i['subcat'] == 'klrnw4w1'):
            pop1anyLS.append(i)
        if(i['subcat'] == '21dr7ggq'):
            pop1anyLSNMG.append(i)
    
    if(i['category'] == 'wkpq95gd'):
        if(i['subcat'] == '21ggx3m1'):
            pop1hundo.append(i)
        if(i['subcat'] == '21ggx3m1'):
            pop1hundoNMG.append(i)
    
    if(i['category'] == 'mkeloznd'):
        pop1pacifist.append(i)
    
    if(i['category'] == 'wdmvjyod'):
        pop1istaria.append(i)
    
    if(i['category'] == '9kv8ww8d'):
        pop1RPR.append(i)
    
    if(i['category'] == 'vdomz5y2'):
        pop1SNESanyNW.append(i)
    
    if(i['category'] == '8241l8m2'):
        pop1SNESany.append(i)
    
    if(i['category'] == 'wkpm95gk'):
        pop1SNESanyNGS.append(i)
    
    if(i['category'] == 'w20go3zk'):
        pop1SNESSR.append(i)
    
    if(i['category'] == 'mkev0xx2'):
        pop1SNES30AP.append(i)
    
    if(i['category'] == '5dw9l1n2'):
        pop1SNESTQoL.append(i)
    
    if(i['category'] == 'wdmxp4ek'):
        pop1SNESEvol.append(i)
    
    if(i['category'] == 'vdo8eo6k'):
        pop1SNESTDC.append(i)
    
    if(i['category'] == 'jdrpw3x2'):
        pop2any.append(i)
    
    if(i['category'] == 'jdz7lxvk'):
        pop3dany.append(i)
    
    if(i['category'] == 'z27zlok0'):
        pop3danyAT.append(i)
    
    if(i['category'] == 'rklv09qd'):
        popanany.append(i)
    
    if(i['category'] == 'zd3xnrdn'):
        if(i['subcat'] == '21d5m3le'):
            popsotACZip.append(i)
        if(i['subcat'] == '5q8pmrld'):
            popsotACZipl.append(i)
        if(i['subcat'] == '5leygm1o'):
            popsotACNMG.append(i)
    
    if(i['category'] == 'xd1v9wd8'):
        if(i['subcat'] == '4qy963l7'):
            popsotanyZip.append(i)
        if(i['subcat'] == 'mln926qp'):
            popsotanyZipl.append(i)
        if(i['subcat'] == '810prjlv'):
            popsotanyNMG.append(i)
        
    if(i['category'] == 'zd3lmz8d'):
        if(i['subcat'] == '013w9krq'):
            popsothundoZip.append(i)
        if(i['subcat'] == 'rqv6v2rl'):
            popsothundoNMG.append(i)
        
    if(i['category'] == '02q0gjky'):
        popsotGBAany.append(i)
        
    if(i['category'] == '02qnm92y'):
        if(i['subcat'] == '0q5mvvlp'):
            popwwanyZip.append(i) 
        if(i['subcat'] == '4lxk5g12'):
            popwwanyZipl.append(i)
        if(i['subcat'] == 'zqo7k51y'):
            popwwanyNMG.append(i)
        
    if(i['category'] == 'wkp550dr'):
        if(i['subcat'] == '013ydrl5'):
            popwwTEZip.append(i)
        if(i['subcat'] == '5lexv6qo'):
            popwwTEZipl.append(i)
        if(i['subcat'] == 'rqvngr16'):
            popwwTENMG.append(i)
        
    if(i['category'] == 'mke9qpjd'):
        if(i['subcat'] == '21gy9d61'):
            popwwACNMG.append(i)
    
    if(i['category'] == '9kvoge82'):
        if(i['subcat'] == 'jqzkge2q'):
            popwwAChZipl.append(i)
        if(i['subcat'] == 'xqk56vnq'):
            popwwAChNMG.append(i)
    
    if(i['category'] == 'wdmx13ok'):
        if(i['subcat'] == 'jq63v07q'):
            popwwAAZip.append(i)
    
    if(i['category'] == '5dw11nkg'):
        if(i['subcat'] == 'p12p0vqx'):
            popt2tanyZip.append(i)
        if(i['subcat'] == '81w2wm14'):
            popt2tanyZipl.append(i)
        if(i['subcat'] == '81p25el7'):
            popt2tanyNMG.append(i)
    
    if(i['category'] == 'n2yvvm2o'):
        if(i['subcat'] == '810y85lv'):
            popt2tAPZip.append(i)
        if(i['subcat'] == '9qjorgl4'):
            popt2tAPZipl.append(i)
        if(i['subcat'] == 'rqvyeyq6'):
            popt2tAPNMG.append(i)
            
    if(i['category'] == 'zd3xvzrd'):
        if(i['subcat'] == '81wwzv51'):
            popt2tPSPUSNMG.append(i)
    
    if(i['category'] == '9d8pn3wk'):
        popclassicany.append(i)
    
    if(i['category'] == 'xd1rv6zk'):
        if(i['subcat'] == 'jqzvpd4l'):
            pop2k8anySt.append(i)
        if(i['subcat'] == 'gq793ndl'):
            pop2k8anyNMG.append(i)
    
    if(i['category'] == 'z276g30d'):
        pop2k8AS.append(i)
    
    if(i['category'] == 'zd30xped'):
        pop2k8Epilougeany.append(i)
    
    if(i['category'] == 'zd366lyd'):
        pop2k8Hacked.append(i)
        
    if(i['category'] == 'xd170vwd'):
        pop2k8anyDupe.append(i)
        
    if(i['category'] == '9kv74gjk'):
        if(i['subcat'] == '21go0xmq'):
            poptfkanyCon.append(i)
        if(i['subcat'] == 'gq7z37y1'):
            poptfkanyEmu.append(i)
    
    if(i['category'] == 'wdm4o54d'):
        if(i['subcat'] == 'zqomxg51'):
            poptfsanySt.append(i)
        if(i['subcat'] == 'jqzxo941'):
            poptfsanyNEG.append(i)
        if(i['subcat'] == '013w9jrq'):
            poptfsanyNMG.append(i)
            
    if(i['category'] == 'w201mqzk'):
        poptfsPSPany.append(i)
            
    if(i['category'] == 'zd3l4vnd'):
        if(i['subcat'] == 'xqkr6ey1'):
            poptfsWiianyNG.append(i)
        if(i['subcat'] == 'gq7nz5y1'):
            poptfsWiianyNGP.append(i)
            
    if(i['category'] == '9kvj3q0k'):
        if(i['subcat'] == '81wr08ml'):
            poptfsDSanyCon.append(i)
        if(i['subcat'] == 'zqo2vn5l'):
            poptfsDSanyEmu.append(i)
    
    if(i['category'] == 'wk67pnxd'):
        mobileHA.append(i)
    
    if(i['category'] == 'mkerym6d'):
        mobileSnFR.append(i)
    
    if(i['category'] == '9kvmp78k'):
        mobileSoT.append(i)
    
    if(i['category'] == 'rklqxvn2'):
        mobileWW.append(i)
    
    if(i['category'] == 'ndxmj7r2'):
        mobileT2T.append(i)
    
    if(i['category'] == 'vdoo0mvd'):
        mobileClassic.append(i)
    
    if(i['category'] == 'w20wpv5d'):
        mobile2k8.append(i)
    
    if(i['category'] == 'wdmwzlo2'):
        mobileTFS.append(i)
    
    if(i['category'] == 'rklxvvqk'):
        if(i['subcat'] == 'jq68rvlm0q503m1p'):
            multiSTanyZip.append(i)
        if(i['subcat'] == 'jq68rvlm4lxj62q2'):
            multiSTanyZipl.append(i)
        if(i['subcat'] == 'jq68rvlm814nw0ld'):
            multiSTanyNMG.append(i)
        if(i['subcat'] == '81w6oml40q503m1p'):
            multiSTCompZip.append(i)
        if(i['subcat'] == '81w6oml44lxj62q2'):
            multiSTCompZipl.append(i)
        if(i['subcat'] == '81w6oml4814nw0ld'):
            multiSTCompNMG.append(i)

def allotPoints(catArray):
    catArray = sorted(catArray, key=lambda d: d['time'])
    points = len(catArray)
    for j in catArray:
        if j['runner'] in runnersArray.keys():
            runnersArray[j['runner']] += points
            points -= 1
        else:
            runnersArray[j['runner']] = points
            points -= 1
            
def calcPoints(catArray, runner):
    catArray = sorted(catArray, key=lambda d: d['time'])
    points = len(catArray)
    runnerpoints = 0
    for j in catArray:
        if j['runner'] == runner:
            runnerpoints+=points
        points-=1
    return runnerpoints

allotPoints(pop1anyDOS)
allotPoints(pop1anyDOSNMG)
allotPoints(pop1anyNES)
allotPoints(pop1anyAppleII)
allotPoints(pop1anyTG16)
allotPoints(pop1anySMS)
allotPoints(pop1anyGBC)
allotPoints(pop1anyAmiga)
allotPoints(pop1anyMD)
allotPoints(pop1anyGenesis)
allotPoints(pop1anyLS)
allotPoints(pop1anyLSNMG)
allotPoints(pop1hundo)
allotPoints(pop1hundoNMG)
allotPoints(pop1pacifist)
allotPoints(pop1istaria)
allotPoints(pop1RPR)
allotPoints(pop1SNESanyNW)
allotPoints(pop1SNESany)
allotPoints(pop1SNESanyNGS)
allotPoints(pop1SNESSR)
allotPoints(pop1SNES30AP)
allotPoints(pop1SNESTQoL)
allotPoints(pop1SNESEvol)
allotPoints(pop1SNESTDC)
allotPoints(pop2any)
allotPoints(pop3dany)
allotPoints(pop3danyAT)
allotPoints(popanany)
allotPoints(popsotACZip)
allotPoints(popsotACZipl)
allotPoints(popsotACNMG)
allotPoints(popsotanyZip)
allotPoints(popsotanyZipl)
allotPoints(popsotanyNMG)
allotPoints(popsothundoZip)
allotPoints(popsothundoNMG)
allotPoints(popsotGBAany)
allotPoints(popwwanyZip)
allotPoints(popwwanyZipl)
allotPoints(popwwanyNMG)
allotPoints(popwwTEZip)
allotPoints(popwwTEZipl)
allotPoints(popwwTENMG)
allotPoints(popwwACNMG)
allotPoints(popwwAChZipl)
allotPoints(popwwAChNMG)
allotPoints(popwwAAZip)
allotPoints(popt2tanyZip)
allotPoints(popt2tanyZipl)
allotPoints(popt2tanyNMG)
allotPoints(popt2tAPZip)
allotPoints(popt2tAPZipl)
allotPoints(popt2tAPNMG)
allotPoints(popt2tPSPUSNMG)
allotPoints(popclassicany)
allotPoints(pop2k8anySt)
allotPoints(pop2k8anyNMG)
allotPoints(pop2k8AS)
allotPoints(pop2k8Epilougeany)
allotPoints(pop2k8Hacked)
allotPoints(pop2k8anyDupe)
allotPoints(poptfkanyCon)
allotPoints(poptfkanyEmu)
allotPoints(poptfsanySt)
allotPoints(poptfsanyNEG)
allotPoints(poptfsanyNMG)
allotPoints(poptfsPSPany)
allotPoints(poptfsWiianyNG)
allotPoints(poptfsWiianyNGP)
allotPoints(poptfsDSanyCon)
allotPoints(poptfsDSanyEmu)
allotPoints(mobileHA)
allotPoints(mobileSnFR)
allotPoints(mobileSoT)
allotPoints(mobileWW)
allotPoints(mobileT2T)
allotPoints(mobileClassic)
allotPoints(mobile2k8)
allotPoints(mobileTFS)
allotPoints(multiSTanyZip)
allotPoints(multiSTanyZipl)
allotPoints(multiSTanyNMG)
allotPoints(multiSTCompZip)
allotPoints(multiSTCompZipl)
allotPoints(multiSTCompNMG)

sorted_array = sorted(runnersArray.items(), key=lambda kv: kv[1], reverse=True)

runnercount = 0;
for i in sorted_array:
    runners_sheet.cell(row=runnercount+4, column=1).value = i[0]
    runners_sheet.cell(row=runnercount+4, column=2).value = i[1]
    runners_sheet.cell(row=runnercount+4, column=3).value = calcPoints(pop1anyDOS, i[0])
    runners_sheet.cell(row=runnercount+4, column=4).value = calcPoints(pop1anyDOSNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=5).value = calcPoints(pop1anyNES, i[0])
    runners_sheet.cell(row=runnercount+4, column=6).value = calcPoints(pop1anyAppleII, i[0])
    runners_sheet.cell(row=runnercount+4, column=7).value = calcPoints(pop1anyTG16, i[0])
    runners_sheet.cell(row=runnercount+4, column=8).value = calcPoints(pop1anyGBC, i[0])
    runners_sheet.cell(row=runnercount+4, column=9).value = calcPoints(pop1anySMS, i[0])
    runners_sheet.cell(row=runnercount+4, column=10).value = calcPoints(pop1anyAmiga, i[0])
    runners_sheet.cell(row=runnercount+4, column=11).value = calcPoints(pop1anyMD, i[0])
    runners_sheet.cell(row=runnercount+4, column=12).value = calcPoints(pop1anyGenesis, i[0])
    runners_sheet.cell(row=runnercount+4, column=13).value = calcPoints(pop1anyLS, i[0])
    runners_sheet.cell(row=runnercount+4, column=14).value = calcPoints(pop1anyLSNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=15).value = calcPoints(pop1hundo, i[0])
    runners_sheet.cell(row=runnercount+4, column=16).value = calcPoints(pop1hundoNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=17).value = calcPoints(pop1pacifist, i[0])
    runners_sheet.cell(row=runnercount+4, column=18).value = calcPoints(pop1istaria, i[0])
    runners_sheet.cell(row=runnercount+4, column=19).value = calcPoints(pop1RPR, i[0])
    runners_sheet.cell(row=runnercount+4, column=20).value = calcPoints(pop1SNESanyNW, i[0])
    runners_sheet.cell(row=runnercount+4, column=21).value = calcPoints(pop1SNESany, i[0])
    runners_sheet.cell(row=runnercount+4, column=22).value = calcPoints(pop1SNESanyNGS, i[0])
    runners_sheet.cell(row=runnercount+4, column=23).value = calcPoints(pop1SNESSR, i[0])
    runners_sheet.cell(row=runnercount+4, column=24).value = calcPoints(pop1SNES30AP, i[0])
    runners_sheet.cell(row=runnercount+4, column=25).value = calcPoints(pop1SNESTQoL, i[0])
    runners_sheet.cell(row=runnercount+4, column=26).value = calcPoints(pop1SNESEvol, i[0])
    runners_sheet.cell(row=runnercount+4, column=27).value = calcPoints(pop1SNESTDC, i[0])
    runners_sheet.cell(row=runnercount+4, column=28).value = calcPoints(pop2any, i[0])
    runners_sheet.cell(row=runnercount+4, column=29).value = calcPoints(pop3dany, i[0])
    runners_sheet.cell(row=runnercount+4, column=30).value = calcPoints(pop3danyAT, i[0])
    runners_sheet.cell(row=runnercount+4, column=31).value = calcPoints(popanany, i[0])
    runners_sheet.cell(row=runnercount+4, column=32).value = calcPoints(popsotanyZip, i[0])
    runners_sheet.cell(row=runnercount+4, column=33).value = calcPoints(popsotanyZipl, i[0])
    runners_sheet.cell(row=runnercount+4, column=34).value = calcPoints(popsotanyNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=35).value = calcPoints(popsotACZip, i[0])
    runners_sheet.cell(row=runnercount+4, column=36).value = calcPoints(popsotACZipl, i[0])
    runners_sheet.cell(row=runnercount+4, column=37).value = calcPoints(popsotACNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=38).value = calcPoints(popsothundoZip, i[0])
    runners_sheet.cell(row=runnercount+4, column=39).value = calcPoints(popsothundoNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=40).value = calcPoints(popsotGBAany, i[0])
    runners_sheet.cell(row=runnercount+4, column=41).value = calcPoints(popwwanyZip, i[0])
    runners_sheet.cell(row=runnercount+4, column=42).value = calcPoints(popwwanyZipl, i[0])
    runners_sheet.cell(row=runnercount+4, column=43).value = calcPoints(popwwanyNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=44).value = calcPoints(popwwTEZip, i[0])
    runners_sheet.cell(row=runnercount+4, column=45).value = calcPoints(popwwTEZipl, i[0])
    runners_sheet.cell(row=runnercount+4, column=46).value = calcPoints(popwwTENMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=47).value = calcPoints(popwwACNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=48).value = calcPoints(popwwAChZipl, i[0])
    runners_sheet.cell(row=runnercount+4, column=49).value = calcPoints(popwwAChNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=50).value = calcPoints(popwwAAZip, i[0])
    runners_sheet.cell(row=runnercount+4, column=51).value = calcPoints(popt2tanyZip, i[0])
    runners_sheet.cell(row=runnercount+4, column=52).value = calcPoints(popt2tanyZipl, i[0])
    runners_sheet.cell(row=runnercount+4, column=53).value = calcPoints(popt2tanyNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=54).value = calcPoints(popt2tAPZip, i[0])
    runners_sheet.cell(row=runnercount+4, column=55).value = calcPoints(popt2tAPZipl, i[0])
    runners_sheet.cell(row=runnercount+4, column=56).value = calcPoints(popt2tAPNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=57).value = calcPoints(popt2tPSPUSNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=58).value = calcPoints(popclassicany, i[0])
    runners_sheet.cell(row=runnercount+4, column=59).value = calcPoints(pop2k8anySt, i[0])
    runners_sheet.cell(row=runnercount+4, column=60).value = calcPoints(pop2k8anyNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=61).value = calcPoints(pop2k8AS, i[0])
    runners_sheet.cell(row=runnercount+4, column=62).value = calcPoints(pop2k8Epilougeany, i[0])
    runners_sheet.cell(row=runnercount+4, column=63).value = calcPoints(pop2k8Hacked, i[0])
    runners_sheet.cell(row=runnercount+4, column=64).value = calcPoints(pop2k8anyDupe, i[0])
    runners_sheet.cell(row=runnercount+4, column=65).value = calcPoints(poptfkanyCon, i[0])
    runners_sheet.cell(row=runnercount+4, column=66).value = calcPoints(poptfkanyEmu, i[0])
    runners_sheet.cell(row=runnercount+4, column=67).value = calcPoints(poptfsanySt, i[0])
    runners_sheet.cell(row=runnercount+4, column=68).value = calcPoints(poptfsanyNEG, i[0])
    runners_sheet.cell(row=runnercount+4, column=69).value = calcPoints(poptfsanyNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=70).value = calcPoints(poptfsPSPany, i[0])
    runners_sheet.cell(row=runnercount+4, column=71).value = calcPoints(poptfsWiianyNG, i[0])
    runners_sheet.cell(row=runnercount+4, column=72).value = calcPoints(poptfsWiianyNGP, i[0])
    runners_sheet.cell(row=runnercount+4, column=73).value = calcPoints(poptfsDSanyCon, i[0])
    runners_sheet.cell(row=runnercount+4, column=74).value = calcPoints(poptfsDSanyEmu, i[0])
    runners_sheet.cell(row=runnercount+4, column=75).value = calcPoints(mobileHA, i[0])
    runners_sheet.cell(row=runnercount+4, column=76).value = calcPoints(mobileSnFR, i[0])
    runners_sheet.cell(row=runnercount+4, column=77).value = calcPoints(mobileSoT, i[0])
    runners_sheet.cell(row=runnercount+4, column=78).value = calcPoints(mobileWW, i[0])
    runners_sheet.cell(row=runnercount+4, column=79).value = calcPoints(mobileT2T, i[0])
    runners_sheet.cell(row=runnercount+4, column=80).value = calcPoints(mobileClassic, i[0])
    runners_sheet.cell(row=runnercount+4, column=81).value = calcPoints(mobile2k8, i[0])
    runners_sheet.cell(row=runnercount+4, column=82).value = calcPoints(mobileTFS, i[0])
    runners_sheet.cell(row=runnercount+4, column=83).value = calcPoints(multiSTanyZip, i[0])
    runners_sheet.cell(row=runnercount+4, column=84).value = calcPoints(multiSTanyZipl, i[0])
    runners_sheet.cell(row=runnercount+4, column=85).value = calcPoints(multiSTanyNMG, i[0])
    runners_sheet.cell(row=runnercount+4, column=86).value = calcPoints(multiSTCompZip, i[0])
    runners_sheet.cell(row=runnercount+4, column=87).value = calcPoints(multiSTCompZipl, i[0])
    runners_sheet.cell(row=runnercount+4, column=88).value = calcPoints(multiSTCompNMG, i[0])
    runnercount+=1;

runners_wb.save('Leaderboard.xlsx')